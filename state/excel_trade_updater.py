"""Excel updater for FastLoop trades.xlsx.

This module provides *in-place* updates for `trades.xlsx` (update original rows)
which is more natural than mutating CSV.

Approach
- The workbook's first row is treated as headers.
- Each trade row is keyed by (condition_id, side, recorded_at_utc).
- When reconciliation finds a newly settled outcome, we update the existing
  row's outcome/settled/pnl_usd/pnl_source/final_pnl_usd columns.

This module is best-effort and should never block trading.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class OpenpyxlRequiredError(RuntimeError):
    pass


def ensure_openpyxl_available() -> None:
    """Raise if openpyxl is unavailable.

    We want a hard signal because Excel reconciliation requires openpyxl.
    """
    if load_workbook is None:
        raise OpenpyxlRequiredError(
            "openpyxl is required for Excel updates. Install with: python -m pip install -U openpyxl"
        )


def _row_key(row: dict[str, Any]) -> str | None:
    cid = str(row.get("condition_id") or "").strip()
    side = str(row.get("side") or "").strip().upper()
    ts = str(row.get("recorded_at_utc") or "").strip()
    if not cid or not side or not ts:
        return None
    return f"{cid}:{side}:{ts}"


def update_trade_row(*, xlsx_path: str | Path, key_row: dict[str, Any], updates: dict[str, Any]) -> bool:
    """Update a single row in trades.xlsx matching key_row. Returns True if updated."""
    ensure_openpyxl_available()

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False

    wb = load_workbook(xlsx_path)
    ws = wb.active
    if ws.max_row < 2:
        return False

    headers = [cell.value for cell in ws[1]]
    header_to_col = {str(h): idx + 1 for idx, h in enumerate(headers) if h is not None}

    key = _row_key(key_row)
    if not key:
        return False

    def row_as_dict(r: int) -> dict[str, Any]:
        data: dict[str, Any] = {}
        for h, col in header_to_col.items():
            data[h] = ws.cell(row=r, column=col).value
        return data

    updated = False
    for r in range(2, ws.max_row + 1):
        existing = row_as_dict(r)
        existing_key = _row_key(existing)
        if existing_key != key:
            continue

        for col_name, val in updates.items():
            if col_name not in header_to_col:
                continue
            ws.cell(row=r, column=header_to_col[col_name]).value = val
        updated = True
        break

    if updated:
        wb.save(xlsx_path)

    return updated

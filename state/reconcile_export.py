"""Trade reconciliation export for the dual-wallet event strategy.

The new flow writes event-level rows rather than legacy preopen action rows.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import csv

from strategy.dual_wallet_models import EventResultSummary, OrderSnapshot, format_operation_timestamp
from state.excel_trade_updater import upsert_event_rows


@dataclass
class TradeRecord:
    recorded_at_utc: str
    recorded_at: str
    run_folder: str
    condition_id: str
    market_id: str | None = None
    event_slug: str | None = None
    question: str | None = None
    window_start_utc: str | None = None
    window_end_utc: str | None = None
    direction: str | None = None
    side: str | None = None
    order_type: str | None = None
    post_only: bool | str | None = None
    shares: float | None = None
    price: float | None = None
    amount_usd: float | None = None
    order_id: str | None = None
    execution_route: str | None = None
    mode: str | None = None
    outcome: str | None = None
    settled: bool = False
    pnl_usd: float | None = None
    pnl_source: str | None = None
    final_pnl_usd: float | None = None


@dataclass
class EventTradeRecord:
    recorded_at_utc: str
    recorded_at: str
    run_folder: str
    event_id: str
    event_name: str
    wallet_id: str
    wallet_name: str
    side: str
    direction: str
    operation: str
    amount_usd: float
    price: float | None
    close_price: float | None
    shares: float | None
    order_id: str | None
    status: str | None
    outcome: str
    event_total_pnl_usd: float
    wallet_pnl_usd: float
    is_profit: bool
    settled_at_utc: str | None


def _iso_now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _append_to_csv(path: Path, rows: list[dict[str, Any]]):
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys())
    file_exists = path.exists()
    with open(path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        if not file_exists:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


def export_dual_wallet_event_to_excel(
    *,
    run_folder: str | Path,
    event_state: Any,
    summary: EventResultSummary,
    dry_run: bool,
    upsert: bool = False,
):
    run_folder = Path(run_folder)
    csv_path = run_folder / "trades.csv"
    xlsx_path = run_folder / "trades.xlsx"

    rows: list[dict[str, Any]] = []
    for snapshot in getattr(event_state, "wallet_orders", {}).values():
        if not isinstance(snapshot, OrderSnapshot):
            continue
        wallet = snapshot.wallet
        wallet_pnl = _safe_float(summary.wallet_pnl_usd.get(wallet.wallet_id, 0.0)) or 0.0
        rows.append(asdict(EventTradeRecord(
            recorded_at_utc=_iso_now_utc(),
            recorded_at=format_operation_timestamp(),
            run_folder=str(run_folder),
            event_id=getattr(event_state, "event_id", ""),
            event_name=getattr(event_state, "event_name", ""),
            wallet_id=wallet.wallet_id,
            wallet_name=wallet.wallet_name,
            side=snapshot.side.value,
            direction=snapshot.side.value,
            operation=snapshot.operation.value,
            amount_usd=float(snapshot.amount_usd or 0.0),
            price=_safe_float(snapshot.price),
            close_price=_safe_float(snapshot.close_price),
            shares=_safe_float(snapshot.shares),
            order_id=snapshot.order_id,
            status=snapshot.status,
            outcome=summary.outcome.value,
            event_total_pnl_usd=float(summary.total_pnl_usd),
            wallet_pnl_usd=float(wallet_pnl),
            is_profit=summary.is_profit,
            settled_at_utc=summary.settled_at.isoformat() if summary.settled_at else None,
        )))

    _append_to_csv(csv_path, rows)
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        return

    headers = list(rows[0].keys()) if rows else []
    if not headers:
        return
    if upsert and xlsx_path.exists():
        upsert_event_rows(xlsx_path=xlsx_path, rows=rows, key_fields=["event_id", "wallet_id", "operation"])
        return
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "trades"
        ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(xlsx_path)
